import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)

    sheet = wb['Sheet1']

    for index in range(2, sheet.max_row + 1):
        cell = sheet.cell(index, 3)
        correct_price = cell.value * 0.9
        correct_price_cell = sheet.cell(index, 4)
        correct_price_cell.value = correct_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=4,
                       min_col=4,
                       max_col=4)

    bar_chart = BarChart()
    bar_chart.add_data(values)
    sheet.add_chart(bar_chart, 'e2')
    wb.save(filename)
    

process_workbook('transactions.xlsx')
